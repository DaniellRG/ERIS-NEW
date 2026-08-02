# -*- coding: utf-8 -*-
"""spreadsheet_generator.py — Crea libros de Excel profesionales con openpyxl.
Soporta: multiples hojas, tablas con formulas (SUM, AVERAGE, IF, etc.),
formato (fuentes, colores, bordes, alineacion), graficos (barras/pastel),
anchos de columna. Crea carpeta automaticamente."""
import json
import os
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

HEADER_FILL = PatternFill(start_color='1A3C6E', end_color='1A3C6E', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center')


def spreadsheet_generator(parameters: dict, player=None) -> str:
    action = parameters.get("action", "create").lower()
    if action == "create":
        return _create_workbook(parameters, player)
    elif action == "list_templates":
        return _list_templates()
    return "Acciones: create (crear Excel), list_templates (ayuda)."


def _create_workbook(parameters: dict, player=None) -> str:
    filename = parameters.get("filename", "")
    output_path_raw = parameters.get("output_path", "")
    sheets_raw = parameters.get("sheets", "")
    if not sheets_raw:
        headers_raw = parameters.get("headers", "[]")
        data_raw = parameters.get("data", "[]")
        if headers_raw != "[]" or data_raw != "[]":
            headers = json.loads(headers_raw) if isinstance(headers_raw, str) else headers_raw
            data = json.loads(data_raw) if isinstance(data_raw, str) else data_raw
            combined = [headers] + data if isinstance(headers, list) else data
            sheets_raw = json.dumps([{"name": parameters.get("title", "Sheet1"), "header_row": True, "data": combined}])
    if not sheets_raw:
        sheets_raw = "[]"

    if not _OPENPYXL_OK:
        return "Error: Falta openpyxl. Instala con: pip install openpyxl"

    try:
        sheets = json.loads(sheets_raw) if isinstance(sheets_raw, str) else sheets_raw
    except json.JSONDecodeError:
        return "Error: sheets debe ser un JSON valido."

    try:
        wb = Workbook()
        wb.remove(wb.active)

        for sheet_data in sheets:
            _add_sheet(wb, sheet_data)

        output_dir = _resolve_output_dir(output_path_raw)
        output_dir.mkdir(parents=True, exist_ok=True)

        if not filename:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Libro_ERIS_{ts}.xlsx"
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        output_path = output_dir / filename
        wb.save(str(output_path))

        result = f"Excel creado: {output_path} ({os.path.getsize(output_path)} bytes, {len(sheets)} hojas)"

        if player:
            player.write_log(result)

        return result

    except Exception as e:
        return f"Error creando Excel: {e}"


def _add_sheet(wb, sheet_data: dict):
    name = sheet_data.get("name", "Hoja1")
    data = sheet_data.get("data", [])
    formulas = sheet_data.get("formulas", [])
    charts = sheet_data.get("charts", [])
    col_widths = sheet_data.get("col_widths", {})
    header_row = sheet_data.get("header_row", True)

    if header_row and not data:
        ws = wb.create_sheet(title=name)
        return

    ws = wb.create_sheet(title=name)
    start_row = 1

    if header_row and data:
        headers = data[0] if isinstance(data[0], list) else list(data[0].keys()) if isinstance(data[0], dict) else []
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        start_row = 2
    else:
        headers = []

    if data:
        body_data = data[1:] if header_row else data
        for row_idx, row_data in enumerate(body_data, start_row):
            if isinstance(row_data, list):
                values = row_data
            elif isinstance(row_data, dict):
                values = [row_data.get(h, "") for h in headers]
            else:
                values = [row_data]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = CENTER
                fmt = _detect_format(val)
                if fmt:
                    cell.number_format = fmt

    last_data_row = start_row - 1 + len(body_data) if data else 0

    for f in formulas:
        _apply_formula(ws, f, last_data_row)

    for c in charts:
        _add_chart_to_sheet(ws, c, data, last_data_row)

    for col_str, width in col_widths.items():
        ws.column_dimensions[col_str].width = width

    if not col_widths and headers:
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(14, len(str(h)) + 4)


def _apply_formula(ws, f: dict, last_data_row: int):
    formula = f.get("formula", "")
    col_str = f.get("col", "")
    row = f.get("row", 0)
    target = f.get("target", "")
    fmt = f.get("number_format", "")
    bold = f.get("bold", False)
    label = f.get("label", "")

    if formula and target:
        if label:
            label_row = row if row else last_data_row + 1
            ws.cell(row=label_row, column=_col_to_int(target), value=label).border = THIN_BORDER
        t_row = row if row else last_data_row + 1
        t_col = _col_to_int(target)
        cell = ws.cell(row=t_row, column=t_col)
        cell.value = formula
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        if fmt:
            cell.number_format = fmt
        if bold:
            cell.font = Font(bold=True)

    if formula and col_str and row:
        cell = ws.cell(row=row, column=_col_to_int(col_str))
        cell.value = formula
        cell.border = THIN_BORDER
        if fmt:
            cell.number_format = fmt
        if bold:
            cell.font = Font(bold=True)


def _add_chart_to_sheet(ws, c: dict, data: list, last_data_row: int):
    chart_type = c.get("type", "bar").lower()
    title = c.get("title", "Grafico")
    categories_col = c.get("categories_col", 1)
    data_col = c.get("data_col", 2)
    data_row_start = c.get("data_row_start", 2)
    data_row_end = c.get("data_row_end", 0) or last_data_row
    anchor = c.get("anchor", "F1")

    if chart_type == "bar":
        chart = BarChart()
    elif chart_type == "pie":
        chart = PieChart()
    elif chart_type == "line":
        chart = LineChart()
    else:
        chart = BarChart()

    chart.type = "col" if chart_type == "bar" else chart_type
    chart.title = title
    chart.style = 10

    data_ref = Reference(ws, min_col=data_col, min_row=1, max_col=data_col, max_row=data_row_end)
    cats = Reference(ws, min_col=categories_col, min_row=data_row_start, max_row=data_row_end)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    if chart_type in ("bar", "line"):
        chart.x_axis.title = "Categorias"
        chart.y_axis.title = "Valores"

    chart.width = c.get("width", 18)
    chart.height = c.get("height", 12)

    ws.add_chart(chart, anchor)


def _resolve_output_dir(path_raw: str) -> Path:
    if path_raw:
        p = Path(path_raw)
        if not p.is_absolute():
            try:
                from actions.path_helper import get_desktop_path
                desk = Path(get_desktop_path())
            except Exception:
                desk = Path.home() / "Desktop"
            p = desk / path_raw
        return p
    try:
        from actions.path_helper import get_desktop_path
        desk = Path(get_desktop_path())
    except Exception:
        desk = Path.home() / "Desktop"
    return desk / "ERIS_Excel"


def _detect_format(val):
    if isinstance(val, int):
        return '#,##0'
    if isinstance(val, float):
        return '#,##0.00'
    return ""


def _col_to_int(col: str) -> int:
    col = col.strip().upper()
    result = 0
    for c in col:
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result


def _list_templates() -> str:
    return (
        "Acciones de spreadsheet_generator:\n\n"
        "1. create: Crear libro de Excel\n"
        "   sheets (str): JSON con hojas y datos:\n"
        '     [{\n'
        '       "name": "Ventas",\n'
        '       "header_row": true,\n'
        '       "data": [\n'
        '         ["Producto","Enero","Febrero"],\n'
        '         ["Laptop",120,135]\n'
        '       ],\n'
        '       "formulas": [\n'
        '         {"formula":"=SUM(B2:B10)","col":"B","row":12,"number_format":"#,##0","bold":true,"label":"Total"},\n'
        '         {"formula":"=AVERAGE(B2:B10)","target":"C2","number_format":"0.0"},\n'
        '         {"formula":"=IF(B2>100,\\"OK\\",\\"Bajo\\")","target":"D2"}\n'
        '       ],\n'
        '       "charts": [\n'
        '         {"type":"bar|pie|line","title":"Ventas","categories_col":1,"data_col":2,"anchor":"F1"}\n'
        '       ],\n'
        '       "col_widths": {"A":20,"B":15}\n'
        '     }]\n'
        "   filename (str): Nombre del .xlsx\n"
        "   output_path (str): Carpeta destino (default: Desktop\\ERIS_Excel)\n\n"
        "FORMULAS SOPORTADAS:\n"
        "  =SUM(B2:B10), =AVERAGE(B2:B10), =IF(condicion,valor_si,valor_no)\n"
        "  =MIN(B2:B10), =MAX(B2:B10), =COUNT(B2:B10), =B2*C2 (operaciones aritmeticas)\n"
        "  =B2/B2, =B2-C2, =B2*C2, cualquier formula de Excel valida\n\n"
        "TIPOS DE GRAFICO:\n"
        "  bar -> grafico de barras\n"
        "  pie -> grafico de pastel\n"
        "  line -> grafico de lineas\n\n"
        "Ejemplo:\n"
        '  spreadsheet_generator action=create sheets=\'[{"name":"Ventas","data":[["Producto","Total"],["A",100]],"formulas":[{"formula":"=SUM(B2:B10)","target":"B11","bold":true}]}]\''
    )
